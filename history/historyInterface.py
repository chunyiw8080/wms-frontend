import sys

import openpyxl
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QHeaderView, QApplication, QTableWidgetItem, QFileDialog
from qfluentwidgets import CardWidget, StrongBodyLabel, TableWidget, SmoothMode, LineEdit, PushButton, InfoBar, ComboBox
from backendRequests.jsonRequests import APIClient
from utils.app_logger import get_logger
from utils.worker import Worker
from config import URL

error_logger = get_logger(logger_name='error_logger', log_file='error.log')

class HistoryInterface(QWidget):
    def __init__(self):
        super().__init__()
        self.setObjectName('HistoryInterface')
        self.setWindowTitle('历史库存记录')
        self.records = []
        self.record_date = ''
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # 顶部按钮组
        card_widget = CardWidget(self)
        buttons_layout = QHBoxLayout(card_widget)

        self.year_label = StrongBodyLabel(self)
        self.year_label.setText('年: ')
        self.year_input = LineEdit(self)
        self.year_input.setFixedWidth(100)

        self.month_label = StrongBodyLabel(self)
        self.month_label.setText('月: ')
        self.month_input = LineEdit(self)
        self.month_input.setFixedWidth(100)

        self.search_btn = PushButton('搜索', self)
        self.search_btn.clicked.connect(self.search_data_by_date)

        self.operations_label = StrongBodyLabel('操作: ', self)
        self.operations_combo = ComboBox(self)
        self.operations_combo.addItems(['------', '导出为excel', '从excel文件导入'])
        self.operations_combo.setFixedWidth(200)
        self.exec_btn = PushButton('执行', self)
        self.exec_btn.clicked.connect(self.exec_operations)

        buttons_layout.addWidget(self.year_label)
        buttons_layout.addWidget(self.year_input)
        buttons_layout.addWidget(self.month_label)
        buttons_layout.addWidget(self.month_input)
        buttons_layout.addWidget(self.search_btn)
        buttons_layout.addStretch(1)
        buttons_layout.addWidget(self.operations_label)
        buttons_layout.addWidget(self.operations_combo)
        buttons_layout.addWidget(self.exec_btn)

        layout.addWidget(card_widget)

        self.table_widget = TableWidget(self)
        self.table_widget.setBorderRadius(8)
        self.table_widget.setBorderVisible(True)
        self.table_widget.verticalHeader().setVisible(False)  # 不显示行号
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_widget.scrollDelagate.verticalSmoothScroll.setSmoothMode(SmoothMode.NO_SMOOTH)

        # 设置表头
        self.table_widget.setColumnCount(11)
        self.table_widget.setHorizontalHeaderLabels(
            ['id', '货品名', '型号', '规格', '分类', '期初单价', '期初数量', '期初总价', '期末单价', '期末数量', '期末总价'])
        layout.addWidget(self.table_widget)

        self.setStyleSheet("HistoryInterface {background-color:white;}")
        self.resize(1280, 760)

    def exec_operations(self):
        index = self.operations_combo.currentIndex()
        match index:
            case 1:
                self.export_to_excel()
            case 2:
                self.import_from_excel()

    def search_data_by_date(self):
        year = self.year_input.text().strip()
        month = self.month_input.text().strip()
        if not HistoryInterface.verify_date_input(year, month):
            InfoBar.warning(title='无效输入', content='年份或月份值无效', parent=self, duration=5000)
        else:
            self.record_date = f'{year}-{month}'
            try:
                url = f'{URL}/history/search?year={year}&month={month}'
                response = Worker.unpack_thread_queue(APIClient.get_request, url)
                if response.get('success') is True:
                    self.records = response.get('data')
                    self.populate_table()
                else:
                    InfoBar.warning(title='没有数据', content='没有对应时间段的历史数据', parent=self, duration=5000)
            except Exception as e:
                error_logger.error(f'historyInterface.search_data_by_date: {e}')

    def populate_table(self):
        """生成表格, 遍历从后端获取的数据并逐行填入到表格中"""
        self.table_widget.setRowCount(len(self.records))
        for row, record_info in enumerate(self.records):
            self.setup_table_row(row, record_info)

    def setup_table_row(self, row: int, record_info: dict):
        data_keys = ['id', 'cargo_name', 'model', 'specification', 'categories', 'starting_price', 'starting_count',
                     'starting_total_price', 'closing_price', 'closing_count', 'closing_total_price']
        # 将多选框绘制到每行的首列
        for col, key in enumerate(data_keys):
            value = record_info.get(key, '')
            # 单元格赋值
            item = QTableWidgetItem(str(value))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
            self.table_widget.setItem(row, col, item)  #注意：第一列是多选框，所以在赋值时从第二列开始

    def export_to_excel(self):
        if self.records is None:
            InfoBar.error(title="导出失败", content='请先输入要导出的记录的日期', parent=self, duration=4000)
            return
        try:
            filename = f'{self.record_date}库存记录'
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = filename
            headers = ['id', '年', '月', '货品名', '型号', '规格', '分类', '期初单价', '期初数量', '期初总价', '期末单价', '期末数量', '期末总价']
            ws.append(headers)
            for record in self.records:
                row = [
                    record['id'],
                    record['year'],
                    record['month'],
                    record['cargo_name'],
                    record['model'],
                    record['specification'],
                    record['categories'],
                    record['starting_price'],
                    record['starting_count'],
                    record['starting_total_price'],
                    record['closing_price'],
                    record['closing_count'],
                    record['closing_total_price'],
                ]
                ws.append(row)
            file_path, _ = QFileDialog.getSaveFileName(self, "保存文件", f"{filename}.xlsx", "Excel Files (*.xlsx)")
            if file_path:
                wb.save(file_path)
                InfoBar.success(title="操作成功", content="已成功导出数据为Excel文件", parent=self, duration=4000)
            else:
                InfoBar.warning(title="操作失败", content="操作被终止", parent=self, duration=4000)
        except Exception as e:
            InfoBar.error(title="操作失败", content=str(e), parent=self, duration=4000)
            error_logger.error(f'historyInterface.export_to_excel: {e}')

    def import_from_excel(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "选择文件", "", "Excel Files (*.xlsx)")
            if not file_path:
                InfoBar.warning("操作取消", "用户终止了导入操作", parent=self, duration=4000)
                return
            else:
                wb = openpyxl.load_workbook(filename=file_path)
                ws = wb.active
                expected_header = ['id', '年', '月', '货品名', '型号', '规格', '分类', '期初单价', '期初数量', '期初总价', '期末单价', '期末数量', '期末总价']
                headers = [cell.value for cell in ws[1]]
                # print(headers)
                if headers != expected_header:
                    InfoBar.error(title="操作失败", content="表头与预期不符", parent=self, duration=4000)
                    return
                else:
                    skipped_num = 0
                    dataset = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        record_id, year, month, cargo_name, model, specification, categories, starting_price, starting_count, starting_total_price, closing_price, closing_count, closing_total_price = row
                        if not all([year, month, cargo_name, model, specification, categories, starting_price, starting_count, starting_total_price, closing_price, closing_count, closing_total_price]):
                            skipped_num += 1
                        else:
                            info = {
                                "year": year,
                                "month": month,
                                "cargo_name": cargo_name,
                                "model": model,
                                "specification": specification,
                                "categories": categories,
                                "starting_price": starting_price,
                                "starting_count": starting_count,
                                "starting_total_price": starting_total_price,
                                "closing_price": closing_price,
                                "closing_count": closing_count,
                                "closing_total_price": closing_total_price,
                            }
                            dataset.append(info)
                    url = f'{URL}/history/import'
                    data = {"dataset": dataset}
                    response = Worker.unpack_thread_queue(APIClient.post_request, url, data)
                    print(response)
                    if response.get('success') is True:
                        succeed = response.get('succeed')
                        failed = response.get('failed')
                        InfoBar.success(title='导入成功', content=f'数据已成功导入到数据库,成功导入{succeed}条, 失败{failed}条', parent=self, duration=5000)
                    else:
                        InfoBar.success(title='导入失败', content='无法导入', parent=self,
                                        duration=5000)
        except Exception as e:
            InfoBar.success(title='导入失败', content=str(e), parent=self, duration=5000)
            error_logger.error(f'historyInterface.import_from_excel: {e}')

    @staticmethod
    def verify_date_input(year: str, month: str):
        if not year or not month:
            return False

        if year.isdigit() and len(year) == 4:
            if month.isdigit() and len(month) in [1, 2]:
                month_int = int(month)
                if 1 <= month_int <= 12:
                    return True
        return False

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = HistoryInterface()
    window.show()
    sys.exit(app.exec())
