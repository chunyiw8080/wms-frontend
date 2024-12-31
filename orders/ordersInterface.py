import math
import sys
from datetime import datetime

import openpyxl
import requests
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QHeaderView, QApplication, QTableWidgetItem, QFileDialog
from qfluentwidgets import CardWidget, StrongBodyLabel, ComboBox, LineEdit, PushButton, setCustomStyleSheet, \
    TableWidget, SmoothMode, PrimaryPushButton, InfoBar, CheckBox

from urllib.parse import urlencode
from utils.worker import Worker
from backendRequests.jsonRequests import APIClient
from utils.custom_styles import ADD_BUTTON_STYLE
from utils.db_utils import load_categories
from config import URL
from orders.ordersDialog import AddOrderDialog, UpdateOrderDialog
from utils.functional_utils import convert_date_to_chinese
from utils.app_logger import get_logger

error_logger = get_logger(logger_name='error_logger', log_file='error.log')


class OrdersInterface(QWidget):
    def __init__(self):
        super().__init__()
        self.setObjectName('OrdersInterface')
        self.setWindowTitle('出入库管理')

        # 后端返回的数据条目集合
        self.orders = []

        # 分页属性
        self.current_page = 1
        self.per_page = 20
        self.total_pages = 1

        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        # 父布局 - 垂直布局
        layout = QVBoxLayout(self)

        # 顶部按钮组
        card_widget = CardWidget(self)
        buttons_layout = QHBoxLayout(card_widget)

        self.conditions_label = StrongBodyLabel(self)
        self.conditions_label.setText('按条件过滤: ')
        self.categoriesComboBox = ComboBox(self)
        self.categoriesComboBox.setFixedWidth(130)
        self.categoriesComboBox.addItem('所有类别')
        self.categoriesComboBox.addItems(load_categories())
        #self.categoriesComboBox.currentIndexChanged.connect(self.on_category_selected)

        self.typeComboBox = ComboBox(self)
        self.typeComboBox.setFixedWidth(130)
        self.typeComboBox.addItems(['全部类型', '出库', '入库'])

        self.statusComboBox = ComboBox(self)
        self.statusComboBox.setFixedWidth(130)
        self.statusComboBox.addItems(['所有状态', '已完成', '待确认', '被取消'])

        self.searchButton1 = PushButton('搜索', self)
        self.searchButton1.clicked.connect(self.search_orders_by_conditions)

        self.search_label = StrongBodyLabel(self)
        self.search_label.setText("根据单号搜索: ")
        self.oreder_number_input = LineEdit(self)
        self.oreder_number_input.setPlaceholderText('输入单号')
        self.oreder_number_input.setFixedWidth(150)

        self.searchButton2 = PushButton('搜索', self)
        self.searchButton2.clicked.connect(self.search_orders_by_id)

        self.operations_label = StrongBodyLabel(self)
        self.operations_label.setText('执行操作: ')
        self.operationsComboBox = ComboBox(self)
        self.operationsComboBox.setFixedWidth(150)
        self.operationsComboBox.addItems(
            ['---------------', '新增记录', '批量删除', '批量编辑', '导出为Excel', '从Excel中导入'])
        self.execBtn = PushButton('执行', self)
        setCustomStyleSheet(self.execBtn, ADD_BUTTON_STYLE, ADD_BUTTON_STYLE)
        self.execBtn.clicked.connect(self.exec_operations)

        buttons_layout.addWidget(self.conditions_label)
        buttons_layout.addWidget(self.categoriesComboBox)
        buttons_layout.addWidget(self.typeComboBox)
        buttons_layout.addWidget(self.statusComboBox)
        buttons_layout.addWidget(self.searchButton1)

        buttons_layout.addStretch(1)

        buttons_layout.addWidget(self.search_label)
        buttons_layout.addWidget(self.oreder_number_input)
        buttons_layout.addWidget(self.searchButton2)

        buttons_layout.addStretch(1)

        buttons_layout.addWidget(self.operations_label)
        buttons_layout.addWidget(self.operationsComboBox)
        buttons_layout.addWidget(self.execBtn)

        layout.addWidget(card_widget)

        # 添加Table
        self.table_widget = TableWidget(self)
        self.table_widget.setBorderRadius(8)
        self.table_widget.setBorderVisible(True)
        self.table_widget.verticalHeader().setVisible(False)  # 不显示行号
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_widget.scrollDelagate.verticalSmoothScroll.setSmoothMode(SmoothMode.NO_SMOOTH)

        # 设置表头
        self.table_widget.setColumnCount(17)
        self.table_widget.setHorizontalHeaderLabels(
            ["", "单号", "类型", "货品id", "货品名称", "型号", "类别", "供应商", "归属项目", "状态", "经办人",
             "提交日期", "审核日期", "单价",  "数量", "规格", "总价"])

        # self.custom_header = CustomHeaderView(orientation=Qt.Orientation.Horizontal, parent=self.table_widget)
        # self.table_widget.setHorizontalHeader(self.custom_header)

        layout.addWidget(self.table_widget)

        # 创建分页布局
        self.paginationWidget = QWidget(self)
        pagination_layout = QHBoxLayout(self.paginationWidget)
        self.prevButton = PrimaryPushButton("上一页", self)
        self.prevButton.clicked.connect(self.load_prev_page)
        self.pageLabel = StrongBodyLabel(self)
        self.nextButton = PrimaryPushButton("下一页", self)
        self.nextButton.clicked.connect(self.load_next_page)

        pagination_layout.addStretch(10)
        pagination_layout.addWidget(self.prevButton)
        pagination_layout.addWidget(self.pageLabel)
        pagination_layout.addWidget(self.nextButton)

        layout.addWidget(self.paginationWidget)

        self.setStyleSheet("OrdersInterface {background-color:white;}")
        self.resize(1280, 760)

        self.data_mapping = {
            '已完成': 'pass',
            '待确认': 'waiting',
            '被取消': 'reject',
            '入库': 'inbound',
            '出库': 'outbound'
        }

    def get_count(self):
        url = URL + '/orders/count'
        try:
            response = Worker.unpack_thread_queue(APIClient.get_request, url)
            if response.get("success") is True:
                return response.get("data")
            else:
                return 0
        except Exception as e:
            error_logger.error(f'ordersInterface.get_count: {e}')

    def load_data(self):
        try:
            url = URL + f'/orders/page/{self.current_page}'
            # 分类选框为全部类别时的分页控制器
            count = self.get_count()
            self.total_pages = math.ceil(count / self.per_page) if count != 0 else 1
            self.update_pagination_controls()
            result = Worker.unpack_thread_queue(APIClient.get_request, url)

            if isinstance(result, (dict, list)):  # 有效 JSON
                if result['success'] is True:  # 有数据
                    self.orders = result['data']
                    self.populate_table()
                    self.update_pagination_controls()
                elif result['success'] is False:
                    InfoBar.warning(title='获取数据失败', content='无数据', parent=self, duration=5000)
            elif isinstance(result, str):  # 错误信息
                InfoBar.error(title='服务器状态异常', content='无法连接到后端服务器', parent=self, duration=10000)
            else:
                print("Unexpected result:", result)
        except Exception as e:
            error_logger.error(f'ordersInterface.load_data: {e}')

    def populate_table(self):
        """生成表格, 遍历从后端获取的数据并逐行填入到表格中"""
        self.table_widget.setRowCount(len(self.orders))
        for row, order_info in enumerate(self.orders):
            self.setup_table_row(row, order_info)
        self.adjust_column_widths()

    def setup_table_row(self, row: int, order_info: dict):
        # 将多选框绘制到每行的首列
        checkbox = CheckBox()
        self.table_widget.setCellWidget(row, 0, checkbox)
        order_info = self.reconstruct_data(order_info)
        # ["", "单号", "类型", "货品id", "货品名称", "型号", "类别", "单价", "供应商", "归属项目", "状态", "经办人", "提交日期", "审核日期", "操作数量"]
        for col, key in enumerate(
                ['order_id', 'order_type', 'cargo_id', 'cargo_name', 'model', 'categories', 'provider',
                 'project', 'status', 'employee_name', 'published_at', 'processed_at', 'price', 'count', 'specification', 'total_price']):
            value = order_info.get(key, '')
            # 单元格赋值
            item = QTableWidgetItem(str(value))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
            self.table_widget.setItem(row, col + 1, item)  #注意：第一列是多选框，所以在赋值时从第二列开始


    def adjust_column_widths(self):
        # 设置第一列（多选框列）为固定宽度
        self.table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table_widget.setColumnWidth(0, 50)  # 第一列宽度固定为 50 像素

        # 设置第二列（order_id 列）为固定宽度
        self.table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.table_widget.setColumnWidth(1, 150)  # 第二列宽度固定为 150 像素

        self.table_widget.horizontalHeader().setSectionResizeMode(14, QHeaderView.ResizeMode.Fixed)
        self.table_widget.setColumnWidth(14, 10)
        self.table_widget.horizontalHeader().setSectionResizeMode(15, QHeaderView.ResizeMode.Fixed)
        self.table_widget.setColumnWidth(15, 10)

        # 让其他列自动调整宽度以适应表格大小
        for col in range(2, self.table_widget.columnCount()):
            self.table_widget.horizontalHeader().setSectionResizeMode(
                col, QHeaderView.ResizeMode.Stretch
            )

    def reconstruct_data(self, data):
        try:
            order_type_mapping = {
                "inbound": "入库",
                "outbound": "出库"
            }
            status_mapping = {
                "pass": "已完成",
                "waiting": "待处理",
                "reject": "已取消"
            }
            newData = {
                "order_id": data['order_id'],
                "order_type": order_type_mapping[data['order_type']],
                "cargo_id": data['cargo_id'],
                "cargo_name": data['cargo_name'],
                "model": data['model'],
                "categories": data['categories'],
                "price": data['price'],
                "provider": data['provider'],
                "project": data['project'],
                "status": status_mapping[data['status']],
                "employee_name": data['employee_name'],
                "published_at": convert_date_to_chinese(data['published_at']),
                "processed_at": convert_date_to_chinese(data['processed_at']),
                "count": data['count'],
                "specification": data['specification'],
                "total_price": data['total_price']
            }
            return newData
        except Exception as e:
            error_logger.error(f'ordersInterface.reconstruct_data: {e}')

    def print_reciept_action(self, order_id):
        """
        点击按钮后，保存订单信息为pdf文件
        :param order_id: 订单id
        :return:
        """
        try:
            url = URL + f'/orders/print/{order_id}'
            # response = APIClient.get_request(url, is_stream=True)
            response = Worker.unpack_thread_queue(APIClient.get_request, url, is_stream=True)

            if isinstance(response, bytes):
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "保存出入库单",
                    f"出入库单-{order_id}.pdf",
                    "PDF Files (*.pdf);;All Files (*)",
                )
                if file_path:  # 用户选择了保存路径
                    if not file_path.endswith(".pdf"):
                        file_path += ".pdf"  # 确保文件扩展名为 .pdf
                else:
                    InfoBar.warning(title='操作取消', content='用户取消了保存操作', parent=self, duration=5000)
                    return

                with open(file_path, "wb") as file:
                    file.write(response)
                InfoBar.success(title='保存成功', content=f'PDF 保存成功: {file_path}', parent=self, duration=5000)
        except Exception as e:
            InfoBar.error(title='保存失败', content=str(e), parent=self, duration=5000)
            error_logger.error(f'ordersInterface.print_reciept_action: {e}')

    def update_pagination_controls(self):
        # print(f'total_pages: {self.total_pages}')
        # 判断分页是否有数据，如果有显示当前是第几页，一共几页，否则显示0页
        if self.total_pages:
            if self.total_pages == 1:
                self.hide_pagination(False)
            else:
                self.hide_pagination(True)
                self.pageLabel.setText(f'第 {self.current_page} 页, 共 {self.total_pages} 页')
                self.prevButton.setEnabled(self.current_page > 1)
                self.nextButton.setEnabled(self.current_page < self.total_pages)
        else:
            self.pageLabel.setText(f'共 0 页')
            self.prevButton.setDisabled(True)
            self.nextButton.setDisabled(True)

    def hide_pagination(self, res):
        self.paginationWidget.setVisible(res)

    def search_orders_by_id(self):
        try:
            order_id = self.oreder_number_input.text()
            if not order_id:
                self.load_data()
                self.populate_table()
                return
            url = URL + f'/orders/{order_id}'
            # response = APIClient.get_request(url)
            response = Worker.unpack_thread_queue(APIClient.get_request, url)
            # print(response)  # 调试：打印响应内容

            if response.get("success") is True:
                self.orders = response.get("data")
                if isinstance(self.orders, dict):  # 单个订单，转为列表
                    self.orders = [self.orders]
                self.populate_table()
            elif response.get("success") is False:
                InfoBar.warning(
                    title='查询失败',
                    content=response.get('message', '未知错误'),
                    parent=self,
                    duration=4000
                )
        except Exception as e:
            InfoBar.error(
                title='查询失败',
                content=str(e),
                parent=self,
                duration=4000
            )
            error_logger.error(f'ordersInterface.search_orders_by_id: {e}')

    def search_orders_by_conditions(self):
        if self.categoriesComboBox.currentIndex() == 0 and self.typeComboBox.currentIndex() == 0 and self.statusComboBox.currentIndex() == 0:
            self.load_data()
        else:
            try:
                base_url = f'{URL}/orders/search?'
                category = self.categoriesComboBox.currentText()
                order_type = self.typeComboBox.currentText()
                order_status = self.statusComboBox.currentText()
                # print(category, self.data_mapping[order_type], order_status)
                query_params = {
                    'i.categories': category if category != '所有类别' else None,  # 忽略空值
                    'o.order_type': self.data_mapping[order_type] if order_type != '全部类型' else None,
                    'o.status': self.data_mapping[order_status] if order_status != '所有状态' else None,
                }
                query_params = {k: v for k, v in query_params.items() if v is not None}
                # 编码查询参数并拼接 URL
                url = base_url + urlencode(query_params)

                # response = APIClient.get_request(url)
                response = Worker.unpack_thread_queue(APIClient.get_request, url)
                if response.get("success") is True:
                    data = response.get('data')
                    self.orders = data
                    self.populate_table()
                    count = len(self.orders)
                    # print(f'search count: {count}')
                    self.total_pages = math.ceil(count / self.per_page)
                    self.update_pagination_controls()
                elif response.get("success") is False:
                    InfoBar.warning(title='没有结果', content='没有与查询条件相匹配的结果', parent=self, duration=5000)
                # 打印结果
            except Exception as e:
                InfoBar.error(title='查询失败', content=str(e), parent=self, duration=5000)
                error_logger.error(f'ordersInterface.search_orders_by_conditions: {e}')

    def get_selected_order_ids(self) -> list:
        """获取并返回多选框被选中的数据的cargo_id"""
        selected_order_ids = []
        for row in range(self.table_widget.rowCount()):
            checkbox = self.table_widget.cellWidget(row, 0)
            if checkbox.isChecked():
                cargo_id = str(self.table_widget.item(row, 1).text())
                selected_order_ids.append(cargo_id)
        return selected_order_ids

    def exec_operations(self):
        index = self.operationsComboBox.currentIndex()
        match index:
            case 1:
                self.add_order()
            case 2:
                self.batch_delete_orders()
            case 3:
                ids = self.get_selected_order_ids()
                for order_id in ids:
                    self.batch_update_orders(order_id)
            case 4:
                self.export_to_excel()
            case 5:
                self.import_from_excel()

    def add_order(self):
        try:
            dialog = AddOrderDialog(parent=self)
            if dialog.exec():
                url = URL + f'/orders/create'
                order_info = dialog.get_order_info()
                # response = APIClient.post_request(url, order_info)
                response = Worker.unpack_thread_queue(APIClient.post_request, url, order_info)
                # print(response)
                if response.get("success") is True:
                    InfoBar.success(title='操作成功', content=response.get("message"), parent=self, duration=4000)
                    self.load_data()
                elif response.get("success") is False:
                    InfoBar.error(title='操作失败', content=response.get("message"), parent=self, duration=4000)
                else:
                    InfoBar.error(title='操作失败', content=response.get("error"), parent=self, duration=4000)
        except Exception as e:
            InfoBar.error(title='系统错误', content=str(e), parent=self, duration=4000)
            error_logger.error(f'ordersInterface.add_order: {e}')

    def batch_delete_orders(self):
        pass

    def batch_update_orders(self, order_id):
        try:
            dialog = UpdateOrderDialog(order_id=order_id, parent=self)
            if dialog.exec():
                new_order_info = dialog.get_order_info()
                url = URL + f'/orders/update/{order_id}'
                # response = APIClient.post_request(url, new_order_info)
                response = Worker.unpack_thread_queue(APIClient.post_request, url, new_order_info)
                if response.get("success") is True:
                    InfoBar.success(title='操作成功', content=response.get("message"), parent=self, duration=4000)
                    self.load_data()
                else:
                    InfoBar.error(title='操作失败', content=response.get("message"), parent=self, duration=4000)
        except Exception as e:
            InfoBar.error(title='操作失败', content=str(e), parent=self, duration=4000)
            error_logger.error(f'ordersInterface.batch_update_orders: {e}')

    def export_to_excel(self):
        try:
            print_data = []
            ids = self.get_selected_order_ids()
            print(ids)
            if len(ids) != 0:
                url = URL + '/orders/batch_query'
                data = {"ids": ids}
                response = Worker.unpack_thread_queue(APIClient.post_request, url, data)
                if response.get('success') is True:
                    print_data = response.get("data")
            else:
                url = URL + '/orders/all'
                response = Worker.unpack_thread_queue(APIClient.get_request, url)
                if response.get("success") is True:
                    print_data = response.get("data")
            # print(print_data)
            if print_data is None:
                InfoBar.warning(title="操作失败", content="无效的数据", parent=self, duration=4000)
                return
            else:
                # print(data)
                date = datetime.today().strftime('%Y-%m-%d')
                filename = f"订单记录-{date}"
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = filename
                headers = ['单号', '类型', '货品id', '货品名称', '型号', '类别', '供应商', '项目', '状态',
                           '经办人', '提交日期', '审核日期', '单价', '数量', '规格', '总价']
                ws.append(headers)
                status_to_chinese = lambda status: {
                    'pass': '完成',
                    'reject': '取消',
                    'waiting': '待处理'
                }.get(status, '未知')
                for order in print_data:
                    row = [
                        order['order_id'],
                        '出库' if order['order_type'] == 'outbound' else '入库',
                        order['cargo_id'],
                        order['cargo_name'],
                        order['model'],
                        order['categories'],
                        order['provider'] if order['provider'] != 'null' else None,
                        order['project'] if order['project'] != 'null' else None,
                        status_to_chinese(order['status']),
                        order['employee_name'],
                        convert_date_to_chinese(order['published_at']),
                        convert_date_to_chinese(order['processed_at']),
                        order['price'],
                        order['count'],
                        order['specification'],
                        order['total_price']
                    ]
                    ws.append(row)
                file_path, _ = QFileDialog.getSaveFileName(self, "保存文件", f"{filename}.xlsx",
                                                           "Excel Files (*.xlsx)")
                if file_path:
                    wb.save(file_path)
                    InfoBar.success(title="操作成功", content="已成功导出数据为Excel文件", parent=self,
                                    duration=4000)
                else:
                    InfoBar.warning(title="操作失败", content="操作被终止", parent=self, duration=4000)
        except Exception as e:
            InfoBar.warning(title="系统错误", content=str(e), parent=self, duration=4000)
            error_logger.error(f'ordersInterface.export_to_excel: {e}')

    def import_from_excel(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "选择文件", "", "Excel Files (*.xlsx)")
            if not file_path:
                InfoBar.warning("操作取消", "用户终止了导入操作", parent=self, duration=4000)
                return
            else:
                wb = openpyxl.load_workbook(filename=file_path)
                ws = wb.active
                expected_header = ['单号', '类型', '货品id', '货品名称', '型号', '类别', '供应商', '项目', '状态',
                               '经办人', '提交日期', '审核日期', '单价', '数量']
                headers = [cell.value for cell in ws[1]]
                print(headers)
                if headers != expected_header:
                    InfoBar.error(title="操作失败", content="表头与预期不符", parent=self, duration=4000)
                    return
                else:
                    skipped, imported = 0, 0
                    dataset = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        order_id, order_type, cargo_id, cargo_name, model, categories, provider, project, status, employee_name, published_at, processed_at, price, count = row
                        if not all([order_id, order_type, cargo_id, cargo_name, model, categories, provider, project, status, employee_name, published_at, processed_at, price, count]):
                            skipped += 1
                        else:
                            info = {
                                'order_id': order_id,
                                'order_type': order_type,
                                'cargo_id': cargo_id,
                                'cargo_name': cargo_name,
                                'model': model,
                                'categories': categories,
                                'provider': provider,
                                'project': project,
                                'status': status,
                                'employee_name': employee_name,
                                'published_at': published_at,
                                'processed_at': processed_at,
                                'price': price,
                                'count': count
                            }
                            dataset.append(info)
                    total_count = len(dataset)
                    url = URL + '/orders/import'
                    json = {
                        "dataset": dataset
                    }
                    # print(f'skipped: {skipped}, dataset: {json}')

                    response = Worker.unpack_thread_queue(APIClient.post_request, url, json)
                    print(response)
                    res = response.get('success')
                    print(res)
                    if res is False:
                        InfoBar.warning(title='导入失败', content=response.get('error'), parent=self, duration=5000)
                    else:
                        if response.get('failed_order_ids') is None:
                            InfoBar.success(title='导入成功', content='导入成功', parent=self, duration=5000)
                        else:
                            failed_order_ids = response.get('failed_order_ids')
                            failed_count = len(failed_order_ids) if failed_order_ids is not None else 0
                            success_count = total_count - failed_count
                            InfoBar.success(title='导入成功', content=f'成功导入了{success_count}条数据, 失败了{failed_count}条数据, 导入失败的订单id为{failed_order_ids}', parent=self, duration=5000)
                        self.load_data()
                        self.populate_table()
        except Exception as e:
            error_logger.error(f'ordersInterface.import_from_excel: {str(e)}')

    def load_prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.load_data()
            self.populate_table()

    def load_next_page(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.load_data()
            self.populate_table()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = OrdersInterface()
    window.show()
    sys.exit(app.exec())
